from flask import Flask, render_template, request, jsonify
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import plotly.utils
import json
import numpy as np
from openpyxl import load_workbook

app = Flask(__name__)

# Custom JSON encoder to handle NumPy arrays
class NumpyJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        if isinstance(obj, np.integer):
            return int(obj)
        if isinstance(obj, np.floating):
            return float(obj)
        return super().default(obj)

# Configure Flask to use our custom JSON encoder
app.json_encoder = NumpyJSONEncoder

# Load and process Excel file
def process_excel_file(file_path):
    # Load workbook to get sheet names
    wb = load_workbook(filename=file_path, read_only=True)
    sheet_names = wb.sheetnames
    
    # Dictionary to store all data
    full_data = []
    
    # Process each sheet
    for sheet_name in sheet_names:
        # Read the Excel sheet
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        # Extract header information (from first two rows)
        headers_row1 = df.iloc[0].dropna().to_dict()
        headers_row2 = df.iloc[1].dropna().to_dict()
        
        # Process data rows (from row 3 onwards)
        for i in range(2, len(df)):
            row = df.iloc[i]
            
            # Skip if no account name
            if pd.isna(row[1]):
                continue
                
            account_name = row[1]
            client_type = row[2] if not pd.isna(row[2]) else ""
            leader = row[3] if not pd.isna(row[3]) else ""
            atl_manager = row[4] if not pd.isna(row[4]) else ""
            
            # Extract all non-empty cells after the management columns (col 8+)
            for col in range(8, len(row)):
                if not pd.isna(row[col]) and row[col] != "":
                    person_name = row[col]
                    
                    # Determine role name from header rows
                    role_category = headers_row1.get(col, "")
                    role_name = headers_row2.get(col, "")
                    
                    # Add to data collection
                    full_data.append({
                        "account": account_name,
                        "client_type": client_type,
                        "leader": leader,
                        "atl_manager": atl_manager,
                        "technology_area": sheet_name,
                        "role_category": role_category,
                        "role": role_name,
                        "person": person_name
                    })
    
    # Convert to DataFrame
    return pd.DataFrame(full_data)

# Home route
@app.route('/')
def index():
    return render_template('index.html')

# Data API endpoints
@app.route('/api/data')
def get_data():
    # Get filter parameters
    account_filter = request.args.get('account', '')
    tech_filter = request.args.get('technology', '')
    role_filter = request.args.get('role', '')
    person_filter = request.args.get('person', '')
    leader_filter = request.args.get('leader', '')
    
    # Apply filters
    filtered_data = data
    if account_filter:
        filtered_data = filtered_data[filtered_data['account'].str.contains(account_filter, case=False)]
    if tech_filter:
        filtered_data = filtered_data[filtered_data['technology_area'] == tech_filter]
    if role_filter:
        filtered_data = filtered_data[filtered_data['role'].str.contains(role_filter, case=False)]
    if person_filter:
        filtered_data = filtered_data[filtered_data['person'].str.contains(person_filter, case=False)]
    if leader_filter:
        filtered_data = filtered_data[filtered_data['leader'].str.contains(leader_filter, case=False)]
    
    # Convert to dict with records orientation
    return jsonify(filtered_data.to_dict(orient='records'))

# Visualization endpoints
@app.route('/api/visualize/accounts')
def visualize_accounts():
    # Count people per account
    account_counts = data.groupby('account').size().reset_index(name='count')
    account_counts = account_counts.sort_values('count', ascending=False)
    
    # Create bar chart
    fig = px.bar(account_counts, x='account', y='count', 
                title='People Deployed Per Account',
                labels={'account': 'Account', 'count': 'Number of People'})
    
    # Customize layout
    fig.update_layout(
        xaxis_tickangle=-45,
        height=600,
        margin=dict(b=100)
    )
    
    # Convert to JSON using plotly's to_json method
    return jsonify(json.loads(fig.to_json()))

@app.route('/api/visualize/technologies')
def visualize_technologies():
    # Count by technology area
    tech_counts = data.groupby(['account', 'technology_area']).size().reset_index(name='count')
    
    # Create grouped bar chart
    fig = px.bar(tech_counts, x='account', y='count', color='technology_area',
                title='People Deployed by Technology Area',
                labels={'account': 'Account', 'count': 'Number of People', 'technology_area': 'Technology Area'})
    
    # Customize layout
    fig.update_layout(
        xaxis_tickangle=-45,
        height=600,
        margin=dict(b=100)
    )
    
    # Convert to JSON using plotly's to_json method
    return jsonify(json.loads(fig.to_json()))

@app.route('/api/filters')
def get_filters():
    # Get unique values for all filter fields
    filters = {
        'accounts': sorted(data['account'].unique().tolist()),
        'technologies': sorted(data['technology_area'].unique().tolist()),
        'roles': sorted(data['role'].unique().tolist()),
        'role_categories': sorted(data['role_category'].unique().tolist()),
        'leaders': sorted(data['leader'].unique().tolist()),
        'managers': sorted(data['atl_manager'].unique().tolist())
    }
    
    return jsonify(filters)

@app.route('/api/account_details/<account_name>')
def account_details(account_name):
    # Filter data for the specific account
    account_data = data[data['account'] == account_name]
    
    # Get technology breakdown
    tech_breakdown = account_data.groupby('technology_area').size().to_dict()
    
    # Get role breakdown
    role_breakdown = account_data.groupby('role').size().to_dict()
    
    # Get list of people
    people_list = account_data[['person', 'role', 'technology_area']].to_dict(orient='records')
    
    return jsonify({
        'account': account_name,
        'total_people': len(account_data),
        'tech_breakdown': tech_breakdown,
        'role_breakdown': role_breakdown,
        'people': people_list
    })

@app.route('/visualize')
def visualize():
    """Additional route for a visualization-focused page"""
    return render_template('visualize.html')

# Add these routes to your Flask app.py file

@app.route('/api/visualize/tech_breakdown')
def visualize_tech_breakdown():
    """Endpoint for tech breakdown pie chart"""
    # Count by technology area
    tech_counts = data.groupby('technology_area').size().reset_index(name='count')
    
    # Create pie chart
    fig = px.pie(tech_counts, names='technology_area', values='count',
                title='Technology Area Distribution',
                labels={'technology_area': 'Technology Area', 'count': 'Number of People'})
    
    # Convert to JSON using plotly's to_json method
    return jsonify(json.loads(fig.to_json()))

@app.route('/api/visualize/role_breakdown')
def visualize_role_breakdown():
    """Endpoint for role breakdown"""
    # Get role parameter (optional)
    role_filter = request.args.get('role', '')
    
    # Filter data if role specified
    filtered_data = data
    if role_filter:
        filtered_data = filtered_data[filtered_data['role'].str.contains(role_filter, case=False)]
    
    # Count by role
    role_counts = filtered_data.groupby('role').size().reset_index(name='count')
    role_counts = role_counts.sort_values('count', ascending=False).head(15)  # Top 15 for readability
    
    # Create bar chart
    fig = px.bar(role_counts, x='role', y='count',
                title='Top Roles by Deployment Count',
                labels={'role': 'Role', 'count': 'Number of People'})
    
    # Customize layout
    fig.update_layout(
        xaxis_tickangle=-45,
        height=500,
        margin=dict(b=100)
    )
    
    # Convert to JSON using plotly's to_json method
    return jsonify(json.loads(fig.to_json()))

@app.route('/api/visualize/account_comparison')
def visualize_account_comparison():
    """Endpoint for comparing accounts"""
    # Get accounts to compare (comma-separated list)
    accounts_str = request.args.get('accounts', '')
    accounts = [acc.strip() for acc in accounts_str.split(',') if acc.strip()]
    
    if not accounts:
        return jsonify({'error': 'No accounts specified for comparison'})
    
    # Filter data to selected accounts
    filtered_data = data[data['account'].isin(accounts)]
    
    # Group by account and technology area
    account_tech_counts = filtered_data.groupby(['account', 'technology_area']).size().reset_index(name='count')
    
    # Create grouped bar chart
    fig = px.bar(account_tech_counts, x='account', y='count', color='technology_area', barmode='group',
                title='Account Comparison by Technology Area',
                labels={'account': 'Account', 'count': 'Number of People', 'technology_area': 'Technology Area'})
    
    # Customize layout
    fig.update_layout(
        xaxis_tickangle=-45,
        height=500,
        margin=dict(b=100)
    )
    
    # Convert to JSON using plotly's to_json method
    return jsonify(json.loads(fig.to_json()))

@app.route('/api/stats')
def get_stats():
    """Endpoint for summary statistics"""
    # Get total accounts
    total_accounts = data['account'].nunique()
    
    # Get total people
    total_people = len(data)
    
    # Get average per account
    avg_per_account = total_people / total_accounts if total_accounts > 0 else 0
    
    # Get top account
    account_counts = data.groupby('account').size()
    top_account = account_counts.idxmax()
    top_account_count = account_counts.max()
    
    # Technology breakdown
    tech_counts = data.groupby('technology_area').size().to_dict()
    
    # Role breakdown (top 10)
    role_counts = data.groupby('role').size().sort_values(ascending=False).head(10).to_dict()
    
    # Return stats
    return jsonify({
        'total_accounts': total_accounts,
        'total_people': total_people,
        'avg_per_account': round(avg_per_account, 1),
        'top_account': top_account,
        'top_account_count': int(top_account_count),
        'tech_counts': tech_counts,
        'role_counts': role_counts
    })

@app.route('/api/export/csv')
def export_csv():
    """Endpoint to export data as CSV"""
    # Get filter parameters
    account_filter = request.args.get('account', '')
    tech_filter = request.args.get('technology', '')
    role_filter = request.args.get('role', '')
    person_filter = request.args.get('person', '')
    
    # Apply filters
    filtered_data = data
    if account_filter:
        filtered_data = filtered_data[filtered_data['account'].str.contains(account_filter, case=False)]
    if tech_filter:
        filtered_data = filtered_data[filtered_data['technology_area'] == tech_filter]
    if role_filter:
        filtered_data = filtered_data[filtered_data['role'].str.contains(role_filter, case=False)]
    if person_filter:
        filtered_data = filtered_data[filtered_data['person'].str.contains(person_filter, case=False)]
    
    # Convert to CSV
    csv_data = filtered_data.to_csv(index=False)
    
    # Create response
    response = app.response_class(
        response=csv_data,
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=account_deployment_data.csv'}
    )
    
    return response

if __name__ == '__main__':
    # Load the data
    try:
        print("Loading Excel file...")
        data = process_excel_file('/Users/mg/Downloads/2025_Technology_Coverage_May2.xlsx')
        print(f"Data loaded successfully. {len(data)} entries found.")
    except Exception as e:
        print(f"Error loading data: {e}")
        data = pd.DataFrame()  # Empty DataFrame as fallback
    
    app.run(debug=True)