from flask import Flask, render_template, request, jsonify
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import plotly.utils
import json
import numpy as np
from openpyxl import load_workbook

app = Flask(__name__)

# Custom JSON encoder to handle NumPy arrays
class NumpyJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        if isinstance(obj, np.integer):
            return int(obj)
        if isinstance(obj, np.floating):
            return float(obj)
        return super().default(obj)

# Configure Flask to use our custom JSON encoder
app.json_encoder = NumpyJSONEncoder

# Load and process Excel file
def process_excel_file(file_path):
    """
    Complete rewrite of the Excel file processing function to ensure accurate personnel counts.
    This version uses a row-scanning approach rather than relying on the structure assumptions.
    """
    import pandas as pd
    from openpyxl import load_workbook
    import re
    
    # Load workbook to get sheet names
    print(f"Loading Excel file: {file_path}")
    wb = load_workbook(filename=file_path, read_only=True)
    sheet_names = wb.sheetnames
    
    # Dictionary to store all data
    full_data = []
    account_personnel_count = {}
    
    # Process each sheet
    for sheet_name in sheet_names:
        print(f"\nProcessing sheet: {sheet_name}")
        
        # Read the entire sheet
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        print(f"Sheet dimensions: {df.shape[0]} rows × {df.shape[1]} columns")
        
        # Find the row that contains "Coverage Name" which should be our header row
        header_row_idx = None
        for i, row in df.iterrows():
            if "Coverage Name" in row.values:
                header_row_idx = i
                print(f"Found header row at index {header_row_idx}")
                break
        
        if header_row_idx is None:
            print(f"Warning: Could not find header row in sheet {sheet_name}, skipping")
            continue
        
        # The row above header_row_idx might contain role categories
        categories_row_idx = header_row_idx - 1 if header_row_idx > 0 else None
        
        # The actual data starts after the header row
        data_start_idx = header_row_idx + 1
        
        # Get column mappings from the header row
        column_mapping = {}
        for col_idx, value in enumerate(df.iloc[header_row_idx]):
            if pd.notna(value) and str(value).strip():
                column_mapping[col_idx] = str(value).strip()
        
        # Find key column indices
        account_col_idx = None
        client_type_col_idx = None
        leader_col_idx = None
        manager_col_idx = None
        
        for col_idx, name in column_mapping.items():
            if name == "Coverage Name":
                account_col_idx = col_idx
            elif name == "Client Sub Type":
                client_type_col_idx = col_idx
            elif name == "Leader":
                leader_col_idx = col_idx
            elif name == "ATL Manager":
                manager_col_idx = col_idx
        
        if account_col_idx is None:
            print(f"Warning: Could not find 'Coverage Name' column in sheet {sheet_name}, skipping")
            continue
        
        # Get role names from header row
        role_names = {}
        for col_idx, name in column_mapping.items():
            # Skip the management columns
            if col_idx <= manager_col_idx:
                continue
            role_names[col_idx] = name
        
        # Get role categories if available
        role_categories = {}
        if categories_row_idx is not None:
            for col_idx, name in role_names.items():
                category = df.iloc[categories_row_idx, col_idx]
                if pd.notna(category) and str(category).strip():
                    role_categories[col_idx] = str(category).strip()
                else:
                    # Look backward to find the nearest category
                    for prev_col in range(col_idx - 1, -1, -1):
                        if prev_col in role_categories:
                            role_categories[col_idx] = role_categories[prev_col]
                            break
        
        # Process each data row
        sheet_personnel_count = 0
        for row_idx in range(data_start_idx, df.shape[0]):
            row = df.iloc[row_idx]
            
            # Skip if account name is missing
            if pd.isna(row[account_col_idx]) or not str(row[account_col_idx]).strip():
                continue
            
            account_name = str(row[account_col_idx]).strip()
            
            # Initialize account count if needed
            if account_name not in account_personnel_count:
                account_personnel_count[account_name] = 0
            
            # Get other management info
            client_type = str(row[client_type_col_idx]).strip() if client_type_col_idx is not None and pd.notna(row[client_type_col_idx]) else ""
            leader = str(row[leader_col_idx]).strip() if leader_col_idx is not None and pd.notna(row[leader_col_idx]) else ""
            manager = str(row[manager_col_idx]).strip() if manager_col_idx is not None and pd.notna(row[manager_col_idx]) else ""
            
            # Process each role column
            row_personnel_count = 0
            for col_idx, role_name in role_names.items():
                cell_value = row[col_idx]
                
                # Skip empty cells or non-person cells
                if pd.isna(cell_value) or not str(cell_value).strip():
                    continue
                    
                person_name = str(cell_value).strip()
                
                # Skip placeholder values (but not actual names)
                placeholder_pattern = r'^(TBD|N/A|None|Select|-)$'
                if re.match(placeholder_pattern, person_name):
                    continue
                
                # Get role category if available
                role_category = role_categories.get(col_idx, "")
                
                # Split multiple names if cell contains commas or slashes
                names = []
                if "," in person_name:
                    names = [name.strip() for name in person_name.split(",") if name.strip()]
                elif "/" in person_name:
                    names = [name.strip() for name in person_name.split("/") if name.strip()]
                else:
                    names = [person_name]
                
                # Add each person
                for name in names:
                    # Skip if after splitting it's a placeholder
                    if re.match(placeholder_pattern, name):
                        continue
                        
                    # Add to data collection
                    full_data.append({
                        "account": account_name,
                        "client_type": client_type,
                        "leader": leader,
                        "atl_manager": manager,
                        "technology_area": sheet_name,
                        "role_category": role_category,
                        "role": role_name,
                        "person": name
                    })
                    
                    row_personnel_count += 1
                    account_personnel_count[account_name] += 1
                    sheet_personnel_count += 1
            
            if row_personnel_count > 0:
                print(f"  Found {row_personnel_count} people for {account_name}")
        
        print(f"  Total in sheet {sheet_name}: {sheet_personnel_count} people")
    
    # Print account counts for verification
    print("\nAccounts by personnel count:")
    sorted_accounts = sorted(account_personnel_count.items(), key=lambda x: x[1], reverse=True)
    for account, count in sorted_accounts:
        print(f"  {account}: {count} people")
    
    print(f"\nTotal personnel across all accounts: {sum(account_personnel_count.values())}")
    print(f"Total records in processed data: {len(full_data)}")
    
    # Convert to DataFrame and return
    return pd.DataFrame(full_data)

# Home route
@app.route('/')
def index():
    return render_template('index.html')

# Data API endpoints
@app.route('/api/data')
def get_data():
    # Get filter parameters
    account_filter = request.args.get('account', '')
    tech_filter = request.args.get('technology', '')
    role_filter = request.args.get('role', '')
    person_filter = request.args.get('person', '')
    leader_filter = request.args.get('leader', '')
    
    # Apply filters
    filtered_data = data
    if account_filter:
        filtered_data = filtered_data[filtered_data['account'].str.contains(account_filter, case=False)]
    if tech_filter:
        filtered_data = filtered_data[filtered_data['technology_area'] == tech_filter]
    if role_filter:
        filtered_data = filtered_data[filtered_data['role'].str.contains(role_filter, case=False)]
    if person_filter:
        filtered_data = filtered_data[filtered_data['person'].str.contains(person_filter, case=False)]
    if leader_filter:
        filtered_data = filtered_data[filtered_data['leader'].str.contains(leader_filter, case=False)]
    
    # Convert to dict with records orientation
    return jsonify(filtered_data.to_dict(orient='records'))

# Visualization endpoints
@app.route('/api/visualize/accounts')
def visualize_accounts():
    # Count people per account
    account_counts = data.groupby('account').size().reset_index(name='count')
    account_counts = account_counts.sort_values('count', ascending=False)
    
    # Create bar chart
    fig = px.bar(account_counts, x='account', y='count', 
                title='People Deployed Per Account',
                labels={'account': 'Account', 'count': 'Number of People'})
    
    # Customize layout
    fig.update_layout(
        xaxis_tickangle=-45,
        height=600,
        margin=dict(b=100)
    )
    
    # Convert to JSON using plotly's to_json method
    return jsonify(json.loads(fig.to_json()))

@app.route('/api/visualize/technologies')
def visualize_technologies():
    # Count by technology area
    tech_counts = data.groupby(['account', 'technology_area']).size().reset_index(name='count')
    
    # Create grouped bar chart
    fig = px.bar(tech_counts, x='account', y='count', color='technology_area',
                title='People Deployed by Technology Area',
                labels={'account': 'Account', 'count': 'Number of People', 'technology_area': 'Technology Area'})
    
    # Customize layout
    fig.update_layout(
        xaxis_tickangle=-45,
        height=600,
        margin=dict(b=100)
    )
    
    # Convert to JSON using plotly's to_json method
    return jsonify(json.loads(fig.to_json()))

@app.route('/api/filters')
def get_filters():
    # Get unique values for all filter fields
    filters = {
        'accounts': sorted(data['account'].unique().tolist()),
        'technologies': sorted(data['technology_area'].unique().tolist()),
        'roles': sorted(data['role'].unique().tolist()),
        'role_categories': sorted(data['role_category'].unique().tolist()),
        'leaders': sorted(data['leader'].unique().tolist()),
        'managers': sorted(data['atl_manager'].unique().tolist())
    }
    
    return jsonify(filters)

@app.route('/api/account_details/<account_name>')
def account_details(account_name):
    # Filter data for the specific account
    account_data = data[data['account'] == account_name]
    
    # Get technology breakdown
    tech_breakdown = account_data.groupby('technology_area').size().to_dict()
    
    # Get role breakdown
    role_breakdown = account_data.groupby('role').size().to_dict()
    
    # Get list of people
    people_list = account_data[['person', 'role', 'technology_area']].to_dict(orient='records')
    
    return jsonify({
        'account': account_name,
        'total_people': len(account_data),
        'tech_breakdown': tech_breakdown,
        'role_breakdown': role_breakdown,
        'people': people_list
    })

@app.route('/visualize')
def visualize():
    """Additional route for a visualization-focused page"""
    return render_template('visualize.html')

# Add these routes to your Flask app.py file

@app.route('/api/visualize/tech_breakdown')
def visualize_tech_breakdown():
    """Endpoint for tech breakdown pie chart"""
    # Count by technology area
    tech_counts = data.groupby('technology_area').size().reset_index(name='count')
    
    # Create pie chart
    fig = px.pie(tech_counts, names='technology_area', values='count',
                title='Technology Area Distribution',
                labels={'technology_area': 'Technology Area', 'count': 'Number of People'})
    
    # Convert to JSON using plotly's to_json method
    return jsonify(json.loads(fig.to_json()))

@app.route('/api/visualize/role_breakdown')
def visualize_role_breakdown():
    """Endpoint for role breakdown"""
    # Get role parameter (optional)
    role_filter = request.args.get('role', '')
    
    # Filter data if role specified
    filtered_data = data
    if role_filter:
        filtered_data = filtered_data[filtered_data['role'].str.contains(role_filter, case=False)]
    
    # Count by role
    role_counts = filtered_data.groupby('role').size().reset_index(name='count')
    role_counts = role_counts.sort_values('count', ascending=False).head(15)  # Top 15 for readability
    
    # Create bar chart
    fig = px.bar(role_counts, x='role', y='count',
                title='Top Roles by Deployment Count',
                labels={'role': 'Role', 'count': 'Number of People'})
    
    # Customize layout
    fig.update_layout(
        xaxis_tickangle=-45,
        height=500,
        margin=dict(b=100)
    )
    
    # Convert to JSON using plotly's to_json method
    return jsonify(json.loads(fig.to_json()))

@app.route('/api/visualize/account_comparison')
def visualize_account_comparison():
    """Endpoint for comparing accounts"""
    # Get accounts to compare (comma-separated list)
    accounts_str = request.args.get('accounts', '')
    accounts = [acc.strip() for acc in accounts_str.split(',') if acc.strip()]
    
    if not accounts:
        return jsonify({'error': 'No accounts specified for comparison'})
    
    # Filter data to selected accounts
    filtered_data = data[data['account'].isin(accounts)]
    
    # Group by account and technology area
    account_tech_counts = filtered_data.groupby(['account', 'technology_area']).size().reset_index(name='count')
    
    # Create grouped bar chart
    fig = px.bar(account_tech_counts, x='account', y='count', color='technology_area', barmode='group',
                title='Account Comparison by Technology Area',
                labels={'account': 'Account', 'count': 'Number of People', 'technology_area': 'Technology Area'})
    
    # Customize layout
    fig.update_layout(
        xaxis_tickangle=-45,
        height=500,
        margin=dict(b=100)
    )
    
    # Convert to JSON using plotly's to_json method
    return jsonify(json.loads(fig.to_json()))

@app.route('/api/stats')
def get_stats():
    """Endpoint for summary statistics"""
    # Get total accounts
    total_accounts = data['account'].nunique()
    
    # Get total people
    total_people = len(data)
    
    # Get average per account
    avg_per_account = total_people / total_accounts if total_accounts > 0 else 0
    
    # Get top account
    account_counts = data.groupby('account').size()
    top_account = account_counts.idxmax()
    top_account_count = account_counts.max()
    
    # Technology breakdown
    tech_counts = data.groupby('technology_area').size().to_dict()
    
    # Role breakdown (top 10)
    role_counts = data.groupby('role').size().sort_values(ascending=False).head(10).to_dict()
    
    # Return stats
    return jsonify({
        'total_accounts': total_accounts,
        'total_people': total_people,
        'avg_per_account': round(avg_per_account, 1),
        'top_account': top_account,
        'top_account_count': int(top_account_count),
        'tech_counts': tech_counts,
        'role_counts': role_counts
    })

@app.route('/api/export/csv')
def export_csv():
    """Endpoint to export data as CSV"""
    # Get filter parameters
    account_filter = request.args.get('account', '')
    tech_filter = request.args.get('technology', '')
    role_filter = request.args.get('role', '')
    person_filter = request.args.get('person', '')
    
    # Apply filters
    filtered_data = data
    if account_filter:
        filtered_data = filtered_data[filtered_data['account'].str.contains(account_filter, case=False)]
    if tech_filter:
        filtered_data = filtered_data[filtered_data['technology_area'] == tech_filter]
    if role_filter:
        filtered_data = filtered_data[filtered_data['role'].str.contains(role_filter, case=False)]
    if person_filter:
        filtered_data = filtered_data[filtered_data['person'].str.contains(person_filter, case=False)]
    
    # Convert to CSV
    csv_data = filtered_data.to_csv(index=False)
    
    # Create response
    response = app.response_class(
        response=csv_data,
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=account_deployment_data.csv'}
    )
    
    return response

if __name__ == '__main__':
    # Load the data
    try:
        print("Loading Excel file...")
        data = process_excel_file('/Users/mg/Downloads/2025_Technology_Coverage_May2.xlsx')
        print(f"Data loaded successfully. {len(data)} entries found.")
    except Exception as e:
        print(f"Error loading data: {e}")
        data = pd.DataFrame()  # Empty DataFrame as fallback
    
    app.run(debug=True)