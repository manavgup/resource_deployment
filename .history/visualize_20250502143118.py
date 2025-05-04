import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import seaborn as sns
from openpyxl import load_workbook

def count_non_empty_cells(row, exclude_cols):
    """Count non-empty cells in a row, excluding specified columns"""
    count = 0
    for idx, val in enumerate(row):
        if idx not in exclude_cols and pd.notna(val) and val != "":
            count += 1
    return count

def analyze_excel_file(file_path):
    """Analyze the Excel file and extract deployment data"""
    # Load the workbook to get sheet names
    wb = load_workbook(filename=file_path, read_only=True)
    sheet_names = wb.sheetnames
    
    # Dictionary to store deployment data
    account_deployment = {}
    
    # Process each sheet
    for sheet_name in sheet_names:
        # Read the sheet, skipping the first two rows (headers)
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, skiprows=2)
        
        # Columns to exclude (management info columns)
        exclude_cols = list(range(8))  # Columns A through H
        
        # Process each row
        for _, row in df.iterrows():
            # Column B (index 1) contains account names
            if pd.notna(row[1]) and isinstance(row[1], str) and row[1].strip() != "":
                account_name = row[1]
                people_deployed = count_non_empty_cells(row, exclude_cols)
                
                # Initialize account if not exists
                if account_name not in account_deployment:
                    account_deployment[account_name] = {
                        'total': 0,
                        'Data': 0,
                        'Automation': 0,
                        'Infrastructure': 0
                    }
                
                # Update deployment counts
                account_deployment[account_name][sheet_name] = people_deployed
                account_deployment[account_name]['total'] += people_deployed
    
    # Convert to DataFrame for easier analysis
    deployment_df = pd.DataFrame(account_deployment).T
    deployment_df = deployment_df.fillna(0).astype(int)
    
    return deployment_df

def create_visualizations(deployment_df):
    """Create visualizations from the deployment data"""
    # Sort by total deployment (descending)
    sorted_df = deployment_df.sort_values('total', ascending=False)
    
    # Create figure with subplots
    plt.figure(figsize=(15, 10))
    
    # 1. Bar chart of top 15 accounts by total deployment
    plt.subplot(2, 1, 1)
    top_15 = sorted_df.head(15)
    
    # Create stacked bar chart
    bottom = np.zeros(len(top_15))
    for col in ['Data', 'Automation', 'Infrastructure']:
        plt.bar(top_15.index, top_15[col], bottom=bottom, label=col)
        bottom += top_15[col].values
    
    plt.title('Top 15 Accounts by Total Deployment', fontsize=14)
    plt.xlabel('Account', fontsize=12)
    plt.ylabel('Number of People Deployed', fontsize=12)
    plt.xticks(rotation=45, ha='right')
    plt.legend()
    plt.grid(axis='y', linestyle='--', alpha=0.7)
    
    # 2. Pie chart of technology distribution
    plt.subplot(2, 2, 3)
    tech_totals = [
        deployment_df['Data'].sum(),
        deployment_df['Automation'].sum(),
        deployment_df['Infrastructure'].sum()
    ]
    plt.pie(tech_totals, labels=['Data', 'Automation', 'Infrastructure'], 
            autopct='%1.1f%%', startangle=90, colors=sns.color_palette('pastel'))
    plt.title('Distribution by Technology Area', fontsize=14)
    
    # 3. Horizontal bar chart for all accounts
    plt.subplot(2, 2, 4)
    plt.barh(range(len(sorted_df)), sorted_df['total'])
    plt.yticks(range(len(sorted_df)), sorted_df.index, fontsize=8)
    plt.title('All Accounts by Total Deployment', fontsize=14)
    plt.xlabel('Number of People Deployed', fontsize=12)
    
    plt.tight_layout()
    plt.savefig('account_deployment_visualization.png', dpi=300, bbox_inches='tight')
    plt.show()
    
    # Print summary statistics
    print("\nSummary Statistics:")
    print(f"Total accounts: {len(deployment_df)}")
    print(f"Total people deployed: {deployment_df['total'].sum()}")
    print("\nDeployment by Technology Area:")
    for col in ['Data', 'Automation', 'Infrastructure']:
        print(f"- {col}: {deployment_df[col].sum()} people ({deployment_df[col].sum() / deployment_df['total'].sum() * 100:.1f}%)")
    
    print("\nTop 5 Accounts by Deployment:")
    for idx, (account, row) in enumerate(sorted_df.head(5).iterrows(), 1):
        print(f"{idx}. {account}: {row['total']} people ({row['Data']} Data, {row['Automation']} Automation, {row['Infrastructure']} Infrastructure)")

def analyze_specific_account(deployment_df, account_name):
    """Analyze a specific account in detail"""
    if account_name in deployment_df.index:
        account_data = deployment_df.loc[account_name]
        
        print(f"\nDetailed Analysis for {account_name}:")
        print(f"Total people deployed: {account_data['total']}")
        print("\nBreakdown by Technology Area:")
        for col in ['Data', 'Automation', 'Infrastructure']:
            print(f"- {col}: {account_data[col]} people ({account_data[col] / account_data['total'] * 100:.1f}%)")
        
        # Create pie chart for this account
        plt.figure(figsize=(8, 6))
        plt.pie([account_data['Data'], account_data['Automation'], account_data['Infrastructure']], 
                labels=['Data', 'Automation', 'Infrastructure'], 
                autopct='%1.1f%%', startangle=90, colors=sns.color_palette('pastel'))
        plt.title(f'Technology Distribution for {account_name}', fontsize=14)
        plt.savefig(f'{account_name.replace(" ", "_")}_analysis.png', dpi=300, bbox_inches='tight')
        plt.show()
    else:
        print(f"Account '{account_name}' not found in the data.")

def main():
    file_path = '2025 Technology Coverage1.xlsx'  # Update this with your file path
    
    print("Analyzing Excel file...")
    deployment_df = analyze_excel_file(file_path)
    
    print("Creating visualizations...")
    create_visualizations(deployment_df)
    
    # Optionally analyze a specific account
    analyze_specific_account(deployment_df, "ROYAL BANK OF CANADA")

if __name__ == "__main__":
    # Set plot style
    plt.style.use('seaborn-v0_8-whitegrid')
    sns.set_palette("pastel")
    
    main()