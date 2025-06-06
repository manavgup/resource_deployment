from flask import Flask, render_template, request, jsonify
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import plotly.utils
import json
import numpy as np
import re
from openpyxl import load_workbook

app = Flask(__name__)

# Custom JSON encoder to handle NumPy arrays
class NumpyJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        if isinstance(obj, np.integer):
            return int(obj)
        if isinstance(obj, np.floating):
            return float(obj)
        return super().default(obj)

# Configure Flask to use our custom JSON encoder
app.json_encoder = NumpyJSONEncoder

# Load and process Excel file
def process_excel_file(file_path):
    """
    Process the Excel file focusing specifically on rows 3 to 56.
    Row indices are 0-based in pandas, so we'll use rows 2 to 55 in the code.
    """
    # Load workbook to get sheet names
    print(f"Loading Excel file: {file_path}")
    wb = load_workbook(filename=file_path, read_only=True)
    sheet_names = wb.sheetnames
    
    # Dictionary to store all data
    full_data = []
    account_personnel_count = {}
    
    # Process each sheet
    for sheet_name in sheet_names:
        print(f"\nProcessing sheet: {sheet_name}")
        
        # Read the entire sheet
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        print(f"Sheet dimensions: {df.shape[0]} rows × {df.shape[1]} columns")
        
        # The first two rows (0 and 1) contain headers
        # Row 0: Role categories
        # Row 1: Specific roles
        role_categories = {}
        for col in range(df.shape[1]):
            if not pd.isna(df.iloc[0, col]) and str(df.iloc[0, col]).strip():
                role_categories[col] = str(df.iloc[0, col]).strip()
        
        roles = {}
        for col in range(df.shape[1]):
            if not pd.isna(df.iloc[1, col]) and str(df.iloc[1, col]).strip():
                roles[col] = str(df.iloc[1, col]).strip()
        
        # Identify key columns
        # We know from the structure that:
        # Column B (index 1): Account name
        # Column C (index 2): Client type
        # Column D (index 3): Leader
        # Column E (index 4): ATL Manager
        account_col_idx = 1
        client_type_col_idx = 2
        leader_col_idx = 3
        manager_col_idx = 4
        
        # Process rows 3 to 56 (indices 2 to 55)
        sheet_personnel_count = 0
        data_rows = range(2, min(56, df.shape[0]))  # Ensure we don't exceed the DataFrame's size
        
        for row_idx in data_rows:
            row = df.iloc[row_idx]
            
            # Skip if account name is missing
            if pd.isna(row[account_col_idx]) or not str(row[account_col_idx]).strip():
                continue
            
            account_name = str(row[account_col_idx]).strip()
            
            # Initialize account count if needed
            if account_name not in account_personnel_count:
                account_personnel_count[account_name] = 0
            
            # Get other management info
            client_type = str(row[client_type_col_idx]).strip() if not pd.isna(row[client_type_col_idx]) else ""
            leader = str(row[leader_col_idx]).strip() if not pd.isna(row[leader_col_idx]) else ""
            manager = str(row[manager_col_idx]).strip() if not pd.isna(row[manager_col_idx]) else ""
            
            # Process each cell starting from column 8 (index 7)
            # This skips the management columns
            row_personnel_count = 0
            for col_idx in range(7, df.shape[1]):
                cell_value = row[col_idx]
                
                # Skip empty cells
                if pd.isna(cell_value) or not str(cell_value).strip():
                    continue
                    
                person_name = str(cell_value).strip()
                
                # Skip placeholder values
                placeholder_pattern = r'^(TBD|N/A|None|Select|-+)$'
                if re.match(placeholder_pattern, person_name):
                    continue
                
                # Get role information
                role_name = ""
                for c in range(col_idx, -1, -1):
                    if c in roles:
                        role_name = roles[c]
                        break
                
                # Get role category
                role_category = ""
                for c in range(col_idx, -1, -1):
                    if c in role_categories:
                        role_category = role_categories[c]
                        break
                
                # Split multiple names if cell contains commas or slashes
                names = []
                if "," in person_name:
                    names = [name.strip() for name in person_name.split(",") if name.strip()]
                elif "/" in person_name:
                    names = [name.strip() for name in person_name.split("/") if name.strip()]
                else:
                    names = [person_name]
                
                # Add each person
                for name in names:
                    # Skip if after splitting it's a placeholder
                    if re.match(placeholder_pattern, name):
                        continue
                        
                    # Add to data collection
                    full_data.append({
                        "account": account_name,
                        "client_type": client_type,
                        "leader": leader,
                        "atl_manager": manager,
                        "technology_area": sheet_name,
                        "role_category": role_category,
                        "role": role_name,
                        "person": name
                    })
                    
                    row_personnel_count += 1
                    account_personnel_count[account_name] += 1
                    sheet_personnel_count += 1
            
            if row_personnel_count > 0:
                print(f"  Row {row_idx+1}: Found {row_personnel_count} people for {account_name}")
        
        print(f"  Total in sheet {sheet_name}: {sheet_personnel_count} people")
    
    # Print account counts for verification
    print("\nAccounts by personnel count:")
    sorted_accounts = sorted(account_personnel_count.items(), key=lambda x: x[1], reverse=True)
    for account, count in sorted_accounts:
        print(f"  {account}: {count} people")
    
    print(f"\nTotal personnel across all accounts: {sum(account_personnel_count.values())}")
    print(f"Total records in processed data: {len(full_data)}")
    
    # Convert to DataFrame and return
    return pd.DataFrame(full_data)

# Home route
@app.route('/')
def index():
    return render_template('index.html')

# Data API endpoints
@app.route('/api/data')
def get_data():
    # Get filter parameters
    account_filter = request.args.get('account', '')
    tech_filter = request.args.get('technology', '')
    role_filter = request.args.get('role', '')
    person_filter = request.args.get('person', '')
    leader_filter = request.args.get('leader', '')
    
    # Apply filters
    filtered_data = data
    if account_filter:
        filtered_data = filtered_data[filtered_data['account'].str.contains(account_filter, case=False)]
    if tech_filter:
        filtered_data = filtered_data[filtered_data['technology_area'] == tech_filter]
    if role_filter:
        filtered_data = filtered_data[filtered_data['role'].str.contains(role_filter, case=False)]
    if person_filter:
        filtered_data = filtered_data[filtered_data['person'].str.contains(person_filter, case=False)]
    if leader_filter:
        filtered_data = filtered_data[filtered_data['leader'].str.contains(leader_filter, case=False)]
    
    # Convert to dict with records orientation
    return jsonify(filtered_data.to_dict(orient='records'))

# Visualization endpoints
@app.route('/api/visualize/accounts')
def visualize_accounts():
    """
    Create a bar chart showing the total number of people per account.
    Fixed to ensure proper counting of all personnel.
    """
    # Count people per account - this is the key part that needs fixing
    account_counts = data.groupby('account').size().reset_index(name='count')
    account_counts = account_counts.sort_values('count', ascending=False)
    
    # For debugging
    print(f"Account count data (top 10):")
    for idx, row in account_counts.head(10).iterrows():
        print(f"  {row['account']}: {row['count']} people")
    
    # Create bar chart
    fig = px.bar(account_counts, x='account', y='count', 
                title='People Deployed Per Account',
                labels={'account': 'Account', 'count': 'Number of People'})
    
    # Customize layout
    fig.update_layout(
        xaxis_tickangle=-45,
        height=600,
        margin=dict(b=100)
    )
    
    # Convert to JSON using plotly's to_json method
    return jsonify(json.loads(fig.to_json()))

@app.route('/api/visualize/technologies')
def visualize_technologies():
    """
    Create a stacked bar chart showing people by technology area per account.
    Fixed to ensure proper counting of all personnel.
    """
    # Count by account and technology area
    tech_counts = data.groupby(['account', 'technology_area']).size().reset_index(name='count')
    
    # For debugging
    print(f"Technology breakdown for top 5 accounts:")
    for account in data['account'].value_counts().head(5).index:
        account_data = tech_counts[tech_counts['account'] == account]
        print(f"  {account}:")
        for idx, row in account_data.iterrows():
            print(f"    {row['technology_area']}: {row['count']} people")
    
    # Create stacked bar chart
    fig = px.bar(tech_counts, x='account', y='count', color='technology_area',
                title='People Deployed by Technology Area',
                labels={'account': 'Account', 'count': 'Number of People', 'technology_area': 'Technology Area'})
    
    # Customize layout
    fig.update_layout(
        xaxis_tickangle=-45,
        height=600,
        margin=dict(b=100)
    )
    
    # Convert to JSON using plotly's to_json method
    return jsonify(json.loads(fig.to_json()))

@app.route('/api/filters')
def get_filters():
    # Get unique values for all filter fields
    filters = {
        'accounts': sorted(data['account'].unique().tolist()),
        'technologies': sorted(data['technology_area'].unique().tolist()),
        'roles': sorted(data['role'].unique().tolist()),
        'role_categories': sorted(data['role_category'].unique().tolist()),
        'leaders': sorted(data['leader'].unique().tolist()),
        'managers': sorted(data['atl_manager'].unique().tolist())
    }
    
    return jsonify(filters)

@app.route('/api/account_details/<account_name>')
def account_details(account_name):
    """
    Get detailed information about a specific account.
    Fixed to ensure proper counting of all personnel.
    """
    # Filter data for the specific account
    account_data = data[data['account'] == account_name]
    
    # Get technology breakdown
    tech_breakdown = account_data.groupby('technology_area').size().to_dict()
    
    # Get role breakdown 
    role_breakdown = account_data.groupby('role').size().to_dict()
    
    # Get list of people
    people_list = account_data[['person', 'role', 'technology_area']].to_dict(orient='records')
    
    # For debugging
    print(f"Account details for {account_name}:")
    print(f"  Total people: {len(account_data)}")
    print(f"  Technology breakdown: {tech_breakdown}")
    print(f"  Number of unique roles: {len(role_breakdown)}")
    print(f"  Number of people records: {len(people_list)}")
    
    return jsonify({
        'account': account_name,
        'total_people': len(account_data),
        'tech_breakdown': tech_breakdown,
        'role_breakdown': role_breakdown,
        'people': people_list
    })

@app.route('/visualize')
def visualize():
    """Additional route for a visualization-focused page"""
    return render_template('visualize.html')

@app.route('/api/visualize/tech_breakdown')
def visualize_tech_breakdown():
    """Endpoint for tech breakdown pie chart"""
    # Count by technology area
    tech_counts = data.groupby('technology_area').size().reset_index(name='count')
    
    # Create pie chart
    fig = px.pie(tech_counts, names='technology_area', values='count',
                title='Technology Area Distribution',
                labels={'technology_area': 'Technology Area', 'count': 'Number of People'})
    
    # Convert to JSON using plotly's to_json method
    return jsonify(json.loads(fig.to_json()))

@app.route('/api/visualize/role_breakdown')
def visualize_role_breakdown():
    """Endpoint for role breakdown"""
    # Get role parameter (optional)
    role_filter = request.args.get('role', '')
    
    # Filter data if role specified
    filtered_data = data
    if role_filter:
        filtered_data = filtered_data[filtered_data['role'].str.contains(role_filter, case=False)]
    
    # Count by role
    role_counts = filtered_data.groupby('role').size().reset_index(name='count')
    role_counts = role_counts.sort_values('count', ascending=False).head(15)  # Top 15 for readability
    
    # Create bar chart
    fig = px.bar(role_counts, x='role', y='count',
                title='Top Roles by Deployment Count',
                labels={'role': 'Role', 'count': 'Number of People'})
    
    # Customize layout
    fig.update_layout(
        xaxis_tickangle=-45,
        height=500,
        margin=dict(b=100)
    )
    
    # Convert to JSON using plotly's to_json method
    return jsonify(json.loads(fig.to_json()))

@app.route('/api/visualize/account_comparison')
def visualize_account_comparison():
    """Endpoint for comparing accounts"""
    # Get accounts to compare (comma-separated list)
    accounts_str = request.args.get('accounts', '')
    accounts = [acc.strip() for acc in accounts_str.split(',') if acc.strip()]
    
    if not accounts:
        return jsonify({'error': 'No accounts specified for comparison'})
    
    # Filter data to selected accounts
    filtered_data = data[data['account'].isin(accounts)]
    
    # Group by account and technology area
    account_tech_counts = filtered_data.groupby(['account', 'technology_area']).size().reset_index(name='count')
    
    # Create grouped bar chart
    fig = px.bar(account_tech_counts, x='account', y='count', color='technology_area', barmode='group',
                title='Account Comparison by Technology Area',
                labels={'account': 'Account', 'count': 'Number of People', 'technology_area': 'Technology Area'})
    
    # Customize layout
    fig.update_layout(
        xaxis_tickangle=-45,
        height=500,
        margin=dict(b=100)
    )
    
    # Convert to JSON using plotly's to_json method
    return jsonify(json.loads(fig.to_json()))

@app.route('/api/stats')
def get_stats():
    """Endpoint for summary statistics"""
    # Get total accounts
    total_accounts = data['account'].nunique()
    
    # Get total people
    total_people = len(data)
    
    # Get average per account
    avg_per_account = total_people / total_accounts if total_accounts > 0 else 0
    
    # Get top account
    account_counts = data.groupby('account').size()
    top_account = account_counts.idxmax()
    top_account_count = account_counts.max()
    
    # Technology breakdown
    tech_counts = data.groupby('technology_area').size().to_dict()
    
    # Role breakdown (top 10)
    role_counts = data.groupby('role').size().sort_values(ascending=False).head(10).to_dict()
    
    # Return stats
    return jsonify({
        'total_accounts': total_accounts,
        'total_people': total_people,
        'avg_per_account': round(avg_per_account, 1),
        'top_account': top_account,
        'top_account_count': int(top_account_count),
        'tech_counts': tech_counts,
        'role_counts': role_counts
    })

@app.route('/api/export/csv')
def export_csv():
    """Endpoint to export data as CSV"""
    # Get filter parameters
    account_filter = request.args.get('account', '')
    tech_filter = request.args.get('technology', '')
    role_filter = request.args.get('role', '')
    person_filter = request.args.get('person', '')
    
    # Apply filters
    filtered_data = data
    if account_filter:
        filtered_data = filtered_data[filtered_data['account'].str.contains(account_filter, case=False)]
    if tech_filter:
        filtered_data = filtered_data[filtered_data['technology_area'] == tech_filter]
    if role_filter:
        filtered_data = filtered_data[filtered_data['role'].str.contains(role_filter, case=False)]
    if person_filter:
        filtered_data = filtered_data[filtered_data['person'].str.contains(person_filter, case=False)]
    
    # Convert to CSV
    csv_data = filtered_data.to_csv(index=False)
    
    # Create response
    response = app.response_class(
        response=csv_data,
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=account_deployment_data.csv'}
    )
    
    return response

# Add these functions to your app.py

def calculate_fte_allocations(data_frame):
    """
    Calculate FTE allocations based on how many accounts each person is assigned to.
    Returns the original dataframe with an added 'allocation' column.
    """
    # Count how many accounts each person appears on
    person_account_counts = data_frame.groupby('person')['account'].nunique()
    
    # Create a dictionary for quick lookup
    allocation_dict = {person: 1.0 / count for person, count in person_account_counts.items()}
    
    # Add allocation column to the dataframe
    data_frame['allocation'] = data_frame['person'].map(allocation_dict)
    
    print(f"Allocation calculation example (first 5 people):")
    for person, count in list(person_account_counts.items())[:5]:
        print(f"  {person}: appears on {count} accounts, allocation = {1.0/count:.2f}")
    
    return data_frame

@app.route('/api/visualize/accounts_fte')
def visualize_accounts_fte():
    """
    Create a bar chart showing the total allocated FTE per account.
    """
    # Calculate FTE allocations if not already done
    if 'allocation' not in data.columns:
        calculate_fte_allocations(data)
    
    # Sum allocations per account
    fte_counts = data.groupby('account')['allocation'].sum().reset_index(name='fte')
    fte_counts = fte_counts.sort_values('fte', ascending=False)
    
    # For debugging
    print(f"FTE count data (top 10):")
    for idx, row in fte_counts.head(10).iterrows():
        print(f"  {row['account']}: {row['fte']:.2f} FTE")
    
    # Create bar chart
    fig = px.bar(fte_counts, x='account', y='fte', 
                title='Allocated FTE Per Account',
                labels={'account': 'Account', 'fte': 'Full-Time Equivalent (FTE)'})
    
    # Customize layout
    fig.update_layout(
        xaxis_tickangle=-45,
        height=600,
        margin=dict(b=100)
    )
    
    # Convert to JSON using plotly's to_json method
    return jsonify(json.loads(fig.to_json()))

@app.route('/api/visualize/technologies_fte')
def visualize_technologies_fte():
    """
    Create a stacked bar chart showing allocated FTE by technology area per account.
    """
    # Calculate FTE allocations if not already done
    if 'allocation' not in data.columns:
        calculate_fte_allocations(data)
    
    # Sum allocations by account and technology area
    fte_counts = data.groupby(['account', 'technology_area'])['allocation'].sum().reset_index(name='fte')
    
    # For debugging
    print(f"Technology FTE breakdown for top 5 accounts:")
    for account in data['account'].value_counts().head(5).index:
        account_data = fte_counts[fte_counts['account'] == account]
        print(f"  {account}:")
        for idx, row in account_data.iterrows():
            print(f"    {row['technology_area']}: {row['fte']:.2f} FTE")
    
    # Create stacked bar chart
    fig = px.bar(fte_counts, x='account', y='fte', color='technology_area',
                title='Allocated FTE by Technology Area',
                labels={'account': 'Account', 'fte': 'Full-Time Equivalent (FTE)', 'technology_area': 'Technology Area'})
    
    # Customize layout
    fig.update_layout(
        xaxis_tickangle=-45,
        height=600,
        margin=dict(b=100)
    )
    
    # Convert to JSON using plotly's to_json method
    return jsonify(json.loads(fig.to_json()))

@app.route('/api/account_details_fte/<account_name>')
def account_details_fte(account_name):
    """
    Get detailed FTE information about a specific account.
    """
    # Calculate FTE allocations if not already done
    if 'allocation' not in data.columns:
        calculate_fte_allocations(data)
    
    # Filter data for the specific account
    account_data = data[data['account'] == account_name]
    
    # Get technology breakdown
    tech_breakdown = account_data.groupby('technology_area')['allocation'].sum().to_dict()
    
    # Get role breakdown 
    role_breakdown = account_data.groupby('role')['allocation'].sum().to_dict()
    
    # Get list of people with their allocations
    people_list = account_data[['person', 'role', 'technology_area', 'allocation']].to_dict(orient='records')
    
    # Get total FTE
    total_fte = account_data['allocation'].sum()
    
    # For debugging
    print(f"Account FTE details for {account_name}:")
    print(f"  Total people: {len(account_data)}")
    print(f"  Total FTE: {total_fte:.2f}")
    print(f"  Technology breakdown: {tech_breakdown}")
    
    return jsonify({
        'account': account_name,
        'total_people': len(account_data),
        'total_fte': total_fte,
        'tech_breakdown': tech_breakdown,
        'role_breakdown': role_breakdown,
        'people': people_list
    })

if __name__ == '__main__':
    # Load the data
    try:
        print("Loading Excel file...")
        data = process_excel_file('/Users/mg/Downloads/2025_Technology_Coverage_May2.xlsx')  # Update this path to match your file location
        print(f"Data loaded successfully. {len(data)} entries found.")
        
        # Extra debug information
        print("\nOverall Statistics:")
        print(f"Total records: {len(data)}")
        print(f"Unique accounts: {data['account'].nunique()}")
        print(f"Unique technology areas: {data['technology_area'].nunique()}")
        print(f"Unique roles: {data['role'].nunique()}")
        
        # Check top accounts
        print("\nTop 10 accounts by personnel count:")
        account_counts = data.groupby('account').size()
        for account, count in account_counts.nlargest(10).items():
            print(f"  {account}: {count}")
        
    except Exception as e:
        print(f"Error loading data: {e}")
        data = pd.DataFrame()  # Empty DataFrame as fallback
    
    app.run(debug=True)